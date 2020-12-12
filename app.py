command = ''
status = False
while True:
    command = input('> ')
    if command == 'start':
        if status:
            print('Car already started..!')
        else:
            status = True
            print('Car start..')
    elif command == 'stop':
        if not status:
            print('Car already stopped..!')
        else:
            status = False
            print('Car is stop..')
    elif command == 'quite':
        break
    else:
        print("Can't understand your command")
        
        #read XL file
        import openpyxl as xl
        from openpyxl.chart import BarChart, Reference


        wb = xl.load_workbook('transaction.xlsx')
        sheet = wb['Sheet1']
        for row in range(2, sheet.max_row + 1):
           cell = sheet.cell(row, 3)
           percentage = cell.value * 0.9
           new_cell = sheet.cell(row, 4)
           new_cell.value = percentage

        values = Reference(sheet,
                  min_row=2,
                  max_row=sheet.max_row,
                  min_col=4,
                  max_col=4)
        chart = BarChart()
        chart.add_data(values)
        sheet.add_chart(chart, 'f2')

        wb.save('transaction.xlsx')
